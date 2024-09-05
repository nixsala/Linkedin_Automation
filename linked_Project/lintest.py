from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import time
import random
import openpyxl
import os

def load_comments(file_name):
    with open(file_name, "r") as f:
        return [line.strip() for line in f.readlines()]

def save_log(log_data, file_name):
    if not os.path.exists(file_name):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Post URL", "Action", "Timestamp"])
        workbook.save(file_name)
    
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active
    for log in log_data:
        sheet.append(log)
    workbook.save(file_name)

def is_post_processed(post_url, file_name):
    if not os.path.exists(file_name):
        return False
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        if post_url in row:
            return True
    return False

driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))

def wait_for_internet():
    while True:
        try:
            driver.execute_script("return window.navigator.onLine;")
            break
        except:
            print("Internet disconnected, waiting...")
            time.sleep(60)  

driver.get("https://www.linkedin.com/login")
time.sleep(2)

email_field = driver.find_element(By.ID, "username")
password_field = driver.find_element(By.ID, "password")

#enter the credentials
email_field.send_keys("your_email@example.com")
password_field.send_keys("your_password")
password_field.send_keys(Keys.RETURN)

time.sleep(5)

driver.get("https://www.linkedin.com/groups/1976445/")
time.sleep(5)

comments = load_comments("comments.txt")

log_file = "action_log.xlsx"
actions_performed = []

while True:
    posts = driver.find_elements(By.CSS_SELECTOR, "div.feed-shared-update-v2")

    for post in posts:
        post_url = post.find_element(By.CSS_SELECTOR, "a.feed-shared-actor__container-link").get_attribute("href")

        if is_post_processed(post_url, log_file):
            continue

        try:
            like_button = post.find_element(By.CSS_SELECTOR, "button[aria-label*='Like']")
            if "active" not in like_button.get_attribute("class"):
                like_button.click()
                actions_performed.append([post_url, "Liked", time.ctime()])
        except Exception as e:
            print(f"Error liking post: {e}")

        try:
            comment_button = post.find_element(By.CSS_SELECTOR, "button[aria-label*='Comment']")
            comment_button.click()
            time.sleep(1)
            comment_box = post.find_element(By.CSS_SELECTOR, "textarea.comments-comment-box__text-editor")
            comment_box.send_keys(random.choice(comments))
            comment_box.send_keys(Keys.RETURN)
            actions_performed.append([post_url, "Commented", time.ctime()])
        except Exception as e:
            print(f"Error commenting on post: {e}")

        try:
            repost_button = post.find_element(By.CSS_SELECTOR, "button[aria-label*='Share']")
            repost_button.click()
            time.sleep(1)
            driver.find_element(By.CSS_SELECTOR, "button[aria-label*='Post']").click()
            actions_performed.append([post_url, "Reposted", time.ctime()])
        except Exception as e:
            print(f"Error reposting post: {e}")

        try:
            follow_button = post.find_element(By.CSS_SELECTOR, "button[aria-label*='Follow']")
            if follow_button.is_displayed():
                follow_button.click()
                actions_performed.append([post_url, "Followed", time.ctime()])
        except Exception as e:
            print(f"Error following user/company: {e}")

        save_log(actions_performed, log_file)
        actions_performed = []

    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(random.randint(5, 10))  

    if len(posts) % 100 == 0:
        print("Pausing for 10 minutes...")
        time.sleep(600)

    wait_for_internet()
